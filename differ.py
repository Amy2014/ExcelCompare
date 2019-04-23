#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import gc
import xlrd

class Rect(object):

    def __init__(self, r):
        self.r = r

    def GetMinRow(self):
        return self.r[0][1]

    def GetMaxRow(self):
        return self.r[1][1]

    def GetMinColumn(self):
        return self.r[0][0]

    def GetMaxColumn(self):
        return self.r[1][0]

class ExcelInfo(object):
    def __init__(self, data, dataRange, sheets_name):
        self.data = data
        self.dataRange = Rect(dataRange)
        # 获取表格sheet名
        self.sheets_name = sheets_name

    def GetMaxColumn(self):
        return self.dataRange.GetMaxColumn()

    def GetMaxRow(self):
        return self.dataRange.GetMaxRow()

    def GetMinColumn(self):
        return self.dataRange.GetMinColumn()

    def GetMinRow(self):
        return self.dataRange.GetMinRow()

    def GetSheetsName(self):
        return self.sheets_name

class ExcelHelper(object):

    @staticmethod
    def OpenExcel(path, index):
        wb =  xlrd.open_workbook(path, on_demand = True)

        # 获取全部sheet名
        sheets_name = wb.sheet_names()

        sheet = wb.sheet_by_name(sheets_name[index])  # 获取当前sheet

        data = []
        minCoordinate, maxCoordinate = [9999999999, 9999999999], [0, 0]
        y = 0
        rows = sheet.get_rows()
        for row in rows:
            _data = []
            x = 0
            for cell in row:
                val = cell.value
                _data.append(val)
                if val:
                    if x < minCoordinate[0]:
                        minCoordinate[0] = x
                    if y < minCoordinate[1]:
                        minCoordinate[1] = y
                    if x > maxCoordinate[0]:
                        maxCoordinate[0] = x
                    if y > maxCoordinate[1]:
                        maxCoordinate[1] = y
                x += 1
            data.append(_data)
            y += 1

        #wb.release_resources()
        return ExcelInfo(data, (minCoordinate, maxCoordinate), sheets_name)

    @staticmethod
    def ColumnIndexFromStr(strCol):
        return column_index_from_string(strCol)

    @staticmethod
    def CoordinateFromStr(strCoor):
        return coordinate_from_string(strCoor)

class ExcelDiffer(object):

    @staticmethod
    def CalcDataRangeIntersection(aRange, bRange):
        aRange = aRange.r
        bRange = bRange.r

        minCoordinate = [0, 0]
        minCoordinate[
            0] = aRange[0][0] if aRange[0][0] > bRange[0][0] else bRange[0][0]
        minCoordinate[
            1] = aRange[0][1] if aRange[0][1] > bRange[0][1] else bRange[0][1]

        maxCoordinate = [0, 0]
        maxCoordinate[
            0] = aRange[1][0] if aRange[1][0] < bRange[1][0] else bRange[1][0]
        maxCoordinate[
            1] = aRange[1][1] if aRange[1][1] < bRange[1][1] else bRange[1][1]

        return (minCoordinate, maxCoordinate)

    @staticmethod
    def IsCoordinateInRect(coordinate, rect):
        return coordinate[0] >= rect[0][0] and coordinate[1] >= rect[0][
            1] and coordinate[0] <= rect[1][0] and coordinate[1] <= rect[1][1]

    @staticmethod
    def GetCoordinate(row, col):
        return get_column_letter(col + 1) + str(row + 1)

    @staticmethod
    def GetColumnLeter(col):
        return get_column_letter(col + 1)

    @staticmethod
    def Diff2(srcExcel, dstExcel, **kwargs):
        '''
        {
            "columns": {
                "new": ["A", "B"],
                "del": ["C"],
            },
            "rows": {
                "new": ["1", "2"],
                "del": ["3"],
            }
            "cels": {
                "A1": ("before", "after"),
            }
        }
        '''
        diffResults = {}

        intersectionRange = ExcelDiffer.CalcDataRangeIntersection(
            srcExcel.dataRange, dstExcel.dataRange)

        diffResults = {}

        maxRows = srcExcel.GetMaxRow() if srcExcel.GetMaxRow(
        ) >= dstExcel.GetMaxRow() else dstExcel.GetMaxRow()
        minRows = srcExcel.GetMinRow() if srcExcel.GetMinRow(
        ) < dstExcel.GetMinRow() else dstExcel.GetMinRow()

        maxCols = srcExcel.GetMaxColumn() if srcExcel.GetMaxColumn(
        ) >= dstExcel.GetMaxColumn() else dstExcel.GetMaxColumn()
        minCols = srcExcel.GetMinColumn() if srcExcel.GetMinColumn(
        ) < dstExcel.GetMinColumn() else dstExcel.GetMinColumn()

        if srcExcel.GetMaxRow() >= dstExcel.GetMaxRow() and srcExcel.GetMinRow() >= dstExcel.GetMinRow():
            diffResults.update({
                "rows": {
                    "new": list(range(minRows+1, intersectionRange[0][1]+1)),
                    "del": list(range(intersectionRange[1][1]+2, maxRows+2)),
                }
            })
        else:
            diffResults.update({
                "rows": {
                    "new": list(range(intersectionRange[1][1]+2, maxRows+2)),
                    "del": list(range(minRows+1, intersectionRange[0][1]+1)),
                }
            })
            

        diffResults["columns"] = {"new":[], "del":[],}
        if srcExcel.GetMaxColumn() >= dstExcel.GetMaxColumn() and srcExcel.GetMinColumn() >= dstExcel.GetMinColumn():
            for delCol in range(minCols, intersectionRange[0][0]):
                diffResults["columns"]["new"].append(get_column_letter(delCol+1))
            for newCol in range(intersectionRange[1][0]+1, maxCols+1):
                diffResults["columns"]["del"].append(get_column_letter(newCol+1))
        else:
            for newCol in range(intersectionRange[1][0]+1, maxCols+1):
                diffResults["columns"]["new"].append(get_column_letter(newCol+1))
            for delCol in range(minCols, intersectionRange[0][0]):
                diffResults["columns"]["del"].append(get_column_letter(delCol+1))

        diffResults["cells"] = {}
        for y in range(intersectionRange[0][1], intersectionRange[1][1]+1):
            res = {}
            for x in range(intersectionRange[0][0], intersectionRange[1][0]+1):
                if srcExcel.data[y][x] == dstExcel.data[y][x]:
                    continue
                res[get_column_letter(x + 1) + str(y + 1)] = (srcExcel.data[y][x], dstExcel.data[y][x])
            if res:
                diffResults["cells"].update(res)

        return diffResults
